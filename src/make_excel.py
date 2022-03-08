import PIL
import openpyxl
import requests
from openpyxl.drawing.image import Image
import connect_databases


def make_excel(excel_name, data_source):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width=7
    ws.cell(row=1, column=1, value='id')
    ws.cell(row=1,column=2,value='name')
    ws.cell(row=1,column=3,value='author')
    ws.cell(row=1, column=4, value='skin_image')
    ws.cell(row=1, column=5, value='preview_image_1')
    ws.cell(row=1, column=6, value='preview_image_2')
    ws.cell(row=1, column=7, value='description')
    ws.cell(row=1, column=8, value='like')
    ws.cell(row=1, column=9, value='comment')
    ws.cell(row=1, column=10, value='share')
    ws.cell(row=1, column=11, value='visit')
    ws.cell(row=1, column=12, value='download')
    ws.cell(row=1, column=13, value='praise')
    ws.cell(row=1, column=14, value='data_source')
    ws.cell(row=1, column=15, value='in_use')
    ws.cell(row=1, column=16, value='size')
    ws.cell(row=1, column=17, value='model')
    ws.cell(row=1, column=18, value='sha256')
    ws.cell(row=1, column=19, value='color_passageway')
    row = 2
    for data in data_source:
        ws.cell(row=row, column=1, value=data[0])
        ws.cell(row=row, column=2, value=data[1])
        ws.cell(row=row, column=3, value=data[2])
        file = '../../pythonProject5/downloadPNG/'+str(data[0])+'download.png'
        img = Image(file)
        img.width = 128
        img.height = 128
        ws.add_image(img, anchor='d' + str(row))
        file = '../../pythonProject5/preview1PNG/'+str(data[0])+'preview1.png'
        try:
            img = Image(file)
            img.width = 128
            img.height = 128
            ws.add_image(img, anchor='e' + str(row))
        except PIL.UnidentifiedImageError as e:
            print(e)
        file = '../../pythonProject5/preview2PNG/' + str(data[0]) + 'preview2.png'
        img = Image(file)
        img.width = 128
        img.height = 128
        ws.add_image(img, anchor='f' + str(row))
        ws.cell(row=row, column=7, value=data[6])
        ws.cell(row=row, column=8, value=data[7])
        ws.cell(row=row, column=9, value=data[8])
        ws.cell(row=row, column=10, value=data[9])
        ws.cell(row=row, column=11, value=data[10])
        ws.cell(row=row, column=12, value=data[11])
        ws.cell(row=row, column=13, value=data[12])
        ws.cell(row=row, column=14, value=data[13])
        ws.cell(row=row, column=15, value=data[15])
        ws.cell(row=row, column=16, value=data[16])
        ws.cell(row=row, column=17, value=data[17])
        ws.cell(row=row, column=18, value=data[18])
        ws.cell(row=row, column=19, value=data[19])

        print(str(row)+" : "+str(data))
        row = row + 1
    wb.save(filename=excel_name + '.xlsx')


def download(url, type_, id):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36 '
    }
    download_response = requests.get(url=url, headers=headers)
    with open('../png_container/' + type_ + id + '.webp', 'wb') as f:
        f.write(download_response.content)
    return '../png_container/' + type_ + id + '.webp'


if __name__ == '__main__':
    db = connect_databases.open_database_mc_skin()
    data = connect_databases.select_from_skin_where_data_source_is_name_mc_order_by_download_desc_limit_2000_remove_duplicate_by_sha256(
        db=db)
    make_excel('name_mc', data)
    connect_databases.close_database(db)
